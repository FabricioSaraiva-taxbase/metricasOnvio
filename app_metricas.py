import streamlit as st
import pandas as pd
import plotly.express as px
import os
import glob
import calendar
from datetime import datetime, timedelta
from openpyxl import load_workbook  # IMPORTANTE: Para salvar sem perder formatação

# --- CONFIGURAÇÃO INICIAL ---
st.set_page_config(page_title="Métricas Taxbase", page_icon="🏢", layout="wide", initial_sidebar_state="collapsed")

# --- 0. GESTÃO DE USUÁRIOS (CREDENCIAIS) ---
CREDENCIAIS = {
    "fabricio": "admin",
    "fernando": "viewer",
    "gustavo": "viewer",
    "admin": "admin"
}
SENHA_PADRAO = "taxbase123"

# --- CONSTANTES VISUAIS ---
COR_PRIMARIA = "#003366"
COR_SECUNDARIA = "#a6a4a4"
COR_CINZA_CLARO = "#F0F2F6"


# --- 1. FUNÇÕES DE BACKEND ---

def autenticar():
    """Gerencia a tela de login simples com suporte a ENTER."""
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False
        st.session_state['user_role'] = None

    if not st.session_state['logged_in']:
        c1, c2, c3 = st.columns([1, 1, 1])
        with c2:
            st.markdown("<br><br>", unsafe_allow_html=True)
            if os.path.exists("logo_taxbase.png"):
                st.image("logo_taxbase.png", use_container_width=True)
            else:
                st.markdown(f"<h2 style='text-align: center; color: {COR_PRIMARIA};'>TAXBASE</h2>",
                            unsafe_allow_html=True)

            st.markdown("<h4 style='text-align: center;'>Acesso Restrito</h4>", unsafe_allow_html=True)

            with st.form("login_form"):
                usuario = st.text_input("Usuário").lower().strip()
                senha = st.text_input("Senha", type="password")
                submit = st.form_submit_button("Entrar", use_container_width=True)

            if submit:
                if usuario in CREDENCIAIS and senha == SENHA_PADRAO:
                    st.session_state['logged_in'] = True
                    st.session_state['user_role'] = CREDENCIAIS[usuario]
                    st.rerun()
                else:
                    st.error("Credenciais inválidas.")
        return False
    return True


def carregar_dados_mes(caminho_arquivo):
    """Lê o CSV, aplica De/Para e corrige fuso horário."""
    if not os.path.exists(caminho_arquivo): return None
    try:
        try:
            df = pd.read_csv(caminho_arquivo, sep=';', encoding='utf-8')
        except:
            df = pd.read_csv(caminho_arquivo, sep=';', encoding='latin1')
    except:
        return None

    # Higienização
    df['Contato_Clean'] = df['Contato'].apply(lambda x: str(x).strip().upper() if pd.notnull(x) else "")

    # Cruzamento com Mestre
    map_path = 'statusContatos.xlsx'
    if os.path.exists(map_path):
        try:
            df_map = pd.read_excel(map_path, engine='openpyxl')
            cols_map = {c.upper(): c for c in df_map.columns}
            if 'NOME DO CONTATO' in cols_map and 'NOME CLIENTE' in cols_map:
                col_nome, col_cli = cols_map['NOME DO CONTATO'], cols_map['NOME CLIENTE']
                df_map_clean = df_map[[col_nome, col_cli]].copy()
                df_map_clean.columns = ['Nome_Map', 'Cliente_Alvo']
                df_map_clean['Nome_Map_Clean'] = df_map_clean['Nome_Map'].apply(
                    lambda x: str(x).strip().upper() if pd.notnull(x) else "")
                df_map_clean = df_map_clean.drop_duplicates(subset=['Nome_Map_Clean'])
                df = pd.merge(df, df_map_clean, left_on='Contato_Clean', right_on='Nome_Map_Clean', how='left')
                df['Cliente_Final'] = df['Cliente_Alvo'].fillna("NÃO IDENTIFICADO")
            else:
                df['Cliente_Final'] = "NÃO IDENTIFICADO"
        except:
            df['Cliente_Final'] = "NÃO IDENTIFICADO"
    else:
        df['Cliente_Final'] = "NÃO IDENTIFICADO"

    # CORREÇÃO CRÍTICA DE DATA
    if 'Data' in df.columns:
        df['Data'] = pd.to_datetime(df['Data'])
        if df['Data'].dt.tz is not None:
            df['Data'] = df['Data'].dt.tz_localize(None)
        df['Dia'] = df['Data'].dt.date
        df = df.sort_values(by='Data', ascending=False)

    return df


def listar_arquivos_por_ano():
    """Organiza os arquivos em dicionário {Ano: [Lista de Arquivos]}."""
    if not os.path.exists("data"): os.makedirs("data")
    arquivos = glob.glob(os.path.join("data", "*.csv"))

    estrutura = {}
    for f in arquivos:
        nome_base = os.path.basename(f).replace(".csv", "")
        # Separa o Ano do resto (Ex: 2026_02 (PARCIAL))
        partes = nome_base.split('_', 1)
        if len(partes) == 2:
            ano = partes[0]
            resto = partes[1]  # Ex: "02 (PARCIAL)" ou "02"

            if ano.isdigit() and len(ano) == 4:
                if ano not in estrutura: estrutura[ano] = []

                # Tenta extrair o mês numérico para ordenação
                mes_num_str = resto[:2]
                try:
                    mes_int = int(mes_num_str)
                except:
                    mes_int = 0

                # --- NOVA LÓGICA DE VISUALIZAÇÃO ---
                # Transforma "02 (PARCIAL)" em "02/26 (PARCIAL)"
                sufixo = resto[2:]  # Pega tudo depois dos 2 primeiros dígitos
                display = f"{mes_num_str}/{ano[2:]}{sufixo}"

                estrutura[ano].append({'caminho': f, 'display': display, 'mes_raw': mes_int})

    for ano in estrutura:
        estrutura[ano] = sorted(estrutura[ano], key=lambda x: x['mes_raw'], reverse=True)

    return dict(sorted(estrutura.items(), reverse=True))


def cadastrar_novo_cliente(nome_contato, nome_cliente):
    """
    Função Feature 3 (CORRIGIDA):
    Usa openpyxl para fazer APPEND na planilha sem sobrescrever formatação.
    """
    map_path = 'statusContatos.xlsx'
    if not os.path.exists(map_path):
        return False, "Arquivo statusContatos.xlsx não encontrado."

    try:
        # Carrega o workbook existente (mantendo estilos)
        wb = load_workbook(map_path)
        ws = wb.active  # Pega a aba ativa

        # Descobre qual coluna é qual
        headers = [cell.value for cell in ws[1]]
        headers_upper = [str(h).upper() if h else "" for h in headers]

        try:
            idx_nome = headers_upper.index('NOME DO CONTATO')
            idx_cliente = headers_upper.index('NOME CLIENTE')
        except ValueError:
            return False, "Colunas 'Nome do Contato' ou 'NOME CLIENTE' não encontradas no cabeçalho."

        # Prepara a nova linha (lista de valores vazios do tamanho do header)
        nova_linha = [None] * len(headers)
        nova_linha[idx_nome] = nome_contato
        nova_linha[idx_cliente] = nome_cliente

        # Adiciona a linha ao final
        ws.append(nova_linha)

        # Salva o arquivo
        wb.save(map_path)
        return True, "Cadastrado com sucesso (Formatação preservada)!"

    except Exception as e:
        return False, str(e)


# --- 2. FRONTEND PRINCIPAL ---

def renderizar_metricas_limpas(df, titulo_contexto, caminho_arquivo_atual=None):
    """Componente reutilizável de dashboard com visual limpo e interativo."""

    # --- ÁREA DE CADASTRO RÁPIDO ---
    pendentes = df[df['Cliente_Final'] == "NÃO IDENTIFICADO"]
    if not pendentes.empty:
        with st.expander(f"🚨 Há {len(pendentes)} atendimentos 'NÃO IDENTIFICADO'. Clique para vincular agora.",
                         expanded=True):
            c_cad1, c_cad2, c_cad3 = st.columns([2, 2, 1])
            with c_cad1:
                lista_nomes = sorted(pendentes['Contato'].unique().astype(str))
                contato_selecionado = st.selectbox("Selecione o Contato", lista_nomes,
                                                   key=f"sel_cont_{titulo_contexto}")
            with c_cad2:
                empresa_input = st.text_input("Nome da Empresa (Cliente)", key=f"inp_cli_{titulo_contexto}",
                                              placeholder="Ex: PADARIA DO JOÃO")
            with c_cad3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("💾 Vincular", key=f"btn_vinc_{titulo_contexto}"):
                    if empresa_input:
                        sucesso, msg = cadastrar_novo_cliente(contato_selecionado, empresa_input.upper())
                        if sucesso:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                    else:
                        st.warning("Digite o nome da empresa.")

    # Espaçamento mínimo
    st.markdown("<div style='margin-top: 5px;'></div>", unsafe_allow_html=True)

    # --- FILTROS ---
    c_f1, c_f2 = st.columns([2, 2])
    with c_f1:
        analistas = ["Todos"] + sorted(list(df['Atendido por'].unique()))
        filtro_analista = st.selectbox(f"Filtrar Analista", analistas, key=f"sel_an_{titulo_contexto}")

    with c_f2:
        clientes_validos = sorted(
            [c for c in df['Cliente_Final'].unique() if c not in ["TAXBASE INTERNO", "IGNORAR", "NÃO IDENTIFICADO"]])
        clientes_opts = ["Todos (Visão Geral)"] + clientes_validos
        filtro_cliente = st.selectbox("Detalhar Empresa Específica", clientes_opts, key=f"sel_cli_{titulo_contexto}")

    # APLICAÇÃO DOS FILTROS
    df_view = df.copy()
    if filtro_analista != "Todos":
        df_view = df_view[df_view['Atendido por'] == filtro_analista]
    if filtro_cliente != "Todos (Visão Geral)":
        df_view = df_view[df_view['Cliente_Final'] == filtro_cliente]

    df_ranking = df_view[~df_view['Cliente_Final'].isin(["TAXBASE INTERNO", "IGNORAR", "NÃO IDENTIFICADO"])]

    # --- KPIS (Margem Reduzida: Removido o <br>) ---
    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
    k1, k2, k3 = st.columns(3)
    k1.metric("Total Geral (Bruto)", len(df_view), help="Total de chamados recebidos (inclui internos)")
    k2.metric("Atendimentos Válidos", len(df_ranking), help="Exclui Taxbase Interno e Ignorar")
    k3.metric("Clientes Únicos", df_ranking['Cliente_Final'].nunique())

    # Linha divisória mais próxima dos KPIs
    st.markdown("<hr style='margin-top: 0px; margin-bottom: 15px;'>", unsafe_allow_html=True)

    # --- GRÁFICOS ---
    c_g1, c_g2 = st.columns([2, 1])
    with c_g1:
        if filtro_cliente != "Todos (Visão Geral)":
            st.markdown(f"#### Volume: {filtro_cliente}")
        else:
            st.markdown("#### Clientes com maior demanda")

        if not df_ranking.empty:
            ranking = df_ranking['Cliente_Final'].value_counts().head(10).reset_index()
            ranking.columns = ['Cliente', 'Volume']

            fig_bar = px.bar(ranking, x='Volume', y='Cliente', orientation='h', text='Volume',
                             color_discrete_sequence=[COR_PRIMARIA])
            fig_bar.update_layout(yaxis={'categoryorder': 'total ascending'}, plot_bgcolor='rgba(0,0,0,0)',
                                  margin=dict(l=0, r=0, t=0, b=0))
            st.plotly_chart(fig_bar, use_container_width=True)

            if filtro_cliente == "Todos (Visão Geral)":
                with st.expander("Ver Ranking Completo (Expandir)"):
                    ranking_full = df_ranking['Cliente_Final'].value_counts().reset_index()
                    ranking_full.columns = ['Cliente', 'Volume']
                    st.dataframe(ranking_full, use_container_width=True, hide_index=True)

    with c_g2:
        st.markdown("#### Atendimentos por Dia")

        usar_mes_completo = st.checkbox("Exibir calendário completo (incluir dias zerados)", value=True,
                                        key=f"chk_{titulo_contexto}")

        if not df_ranking.empty:
            min_date = df_ranking['Dia'].min()
            max_date = df_ranking['Dia'].max()

            if usar_mes_completo:
                start_date = min_date.replace(day=1)
                last_day_num = calendar.monthrange(max_date.year, max_date.month)[1]
                end_date = max_date.replace(day=last_day_num)
                full_range = pd.date_range(start=start_date, end=end_date)

                timeline = df_ranking.groupby('Dia').size()
                timeline.index = pd.to_datetime(timeline.index)
                timeline = timeline.reindex(full_range, fill_value=0).reset_index()
                timeline.columns = ['Dia', 'Volume']
                timeline['Dia'] = timeline['Dia'].dt.date
            else:
                timeline = df_ranking.groupby('Dia').size().reset_index(name='Volume')

            dias_map = {0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira', 3: 'Quinta-feira',
                        4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'}
            timeline['Temp_Date'] = pd.to_datetime(timeline['Dia'])
            timeline['Dia da Semana'] = timeline['Temp_Date'].dt.dayofweek.map(dias_map)

            fig_line = px.line(timeline, x='Dia', y='Volume', markers=True,
                               hover_data={'Dia': '|%d/%m/%Y', 'Volume': True, 'Dia da Semana': True,
                                           'Temp_Date': False})

            fig_line.update_traces(line_color=COR_PRIMARIA, line_width=3)
            fig_line.update_layout(plot_bgcolor='rgba(0,0,0,0)', margin=dict(l=0, r=0, t=0, b=0))
            fig_line.update_yaxes(rangemode="tozero", dtick=1 if timeline['Volume'].max() < 20 else None)
            st.plotly_chart(fig_line, use_container_width=True)

    st.markdown("### 📋 Detalhamento")
    st.dataframe(df_ranking[['Data', 'Cliente_Final', 'Contato', 'Atendido por', 'Status']], use_container_width=True,
                 hide_index=True)

    # --- OPÇÕES DE ADMINISTRADOR (LISTA ÚNICA) ---
    if st.session_state.get('user_role') == 'admin' and caminho_arquivo_atual:
        st.markdown("<br>", unsafe_allow_html=True)

        with st.expander("🛠️ Opções de Administrador"):  # Título ajustado

            # 1. RENOMEAR
            st.markdown("##### ✏️ 1. Renomear Mês")
            col_ren1, col_ren2 = st.columns([3, 1])
            with col_ren1:
                nome_atual_arquivo = os.path.basename(caminho_arquivo_atual).replace(".csv", "")
                novo_nome = st.text_input("Novo nome do arquivo:", value=nome_atual_arquivo,
                                          key=f"rn_in_{titulo_contexto}", label_visibility="collapsed")
            with col_ren2:
                if st.button("Renomear", key=f"rn_bt_{titulo_contexto}"):
                    if novo_nome and novo_nome != nome_atual_arquivo:
                        novo_caminho = os.path.join("data", novo_nome + ".csv")
                        try:
                            os.rename(caminho_arquivo_atual, novo_caminho)
                            st.success("Renomeado!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro: {e}")

            st.markdown("---")

            # 2. SUBSTITUIR
            st.markdown("##### 📂 2. Substituir Arquivo (CSV)")
            col_up1, col_up2 = st.columns([3, 1])
            with col_up1:
                up_sub = st.file_uploader("Upload do novo arquivo", type=['csv'], key=f"up_sub_{titulo_contexto}",
                                          label_visibility="collapsed")
            with col_up2:
                if up_sub and st.button("Substituir", key=f"bt_sub_{titulo_contexto}"):
                    with open(caminho_arquivo_atual, "wb") as f:
                        f.write(up_sub.getbuffer())
                    st.success("Atualizado!")
                    st.rerun()

            st.markdown("---")

            # 3. EXCLUIR
            st.markdown("##### 🗑️ 3. Zona de Perigo")
            col_del1, col_del2 = st.columns([3, 1])
            with col_del1:
                check_del = st.checkbox("Confirmar exclusão definitiva deste mês", key=f"chk_del_{titulo_contexto}")
            with col_del2:
                if st.button("Excluir Mês", type="primary", key=f"bt_del_{titulo_contexto}"):
                    if check_del:
                        try:
                            os.remove(caminho_arquivo_atual)
                            st.success("Excluído!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro: {e}")
                    else:
                        st.warning("Marque a caixa ao lado.")


def main_app():
    # --- CSS BLINDADO V5 (REVISADO) ---
    st.markdown(f"""
        <style>
        /* 1. LAYOUT GERAL */
        .block-container {{ padding-top: 3rem; padding-bottom: 0rem; }}
        section[data-testid="stSidebar"] {{ display: none; }}

        /* 2. TIPOGRAFIA */
        h1, h2, h3, h4 {{ color: {COR_PRIMARIA}; font-family: 'Segoe UI', sans-serif; }}
        div[data-testid="stMetricValue"] {{ color: {COR_PRIMARIA}; font-weight: bold; }}

        /* 3. CORREÇÃO DAS ABAS - REMOÇÃO DE BORDAS */
        .stTabs [data-baseweb="tab-list"] {{
            border-bottom: none !important;
            gap: 8px;
            padding-bottom: 0px !important;
        }}
        .stTabs [data-baseweb="tab"] {{
            border: none !important;
            background-color: transparent !important;
            box-shadow: none !important;
            outline: none !important;
            color: {COR_SECUNDARIA};
        }}
        /* Elemento "fantasma" que causa a linha vermelha */
        div[data-baseweb="tab-highlight"] {{
            display: none !important;
        }}
        /* Aba Ativa */
        .stTabs [aria-selected="true"] {{
            border-bottom: 3px solid {COR_PRIMARIA} !important;
            color: {COR_PRIMARIA} !important;
            font-weight: bold !important;
        }}
        button[data-baseweb="tab"][aria-selected="true"] p {{
            color: {COR_PRIMARIA} !important;
        }}

        /* 4. INPUTS AZUIS */
        div[data-baseweb="select"]:focus-within > div, 
        div[data-baseweb="input"]:focus-within > div {{
            border-color: {COR_PRIMARIA} !important;
            box-shadow: 0 0 0 1px {COR_PRIMARIA} !important;
        }}
        input {{ caret-color: {COR_PRIMARIA} !important; }}
        div[data-baseweb="checkbox"] span[aria-checked="true"] {{
            background-color: {COR_PRIMARIA} !important;
            border-color: {COR_PRIMARIA} !important;
        }}

        /* 5. BOTÃO VOLTAR */
        .hub-btn {{ position: fixed; top: 130px; left: 10px; z-index: 9999; }}
        </style>
    """, unsafe_allow_html=True)

    # --- BOTÃO VOLTAR AO HUB ---
    c_hub, _ = st.columns([1, 10])
    with c_hub:
        if st.button("⬅️ Hub", help="Voltar para o Taxbase Hub"):
            st.session_state['logged_in'] = False
            st.rerun()

    # --- HEADER ---
    col_spacer_L, col_center, col_spacer_R = st.columns([1, 2, 1])
    with col_center:
        if os.path.exists("logo_taxbase.png"):
            st.markdown(f"""<div style="display: flex; justify-content: center;"></div>""", unsafe_allow_html=True)

    c1, c2, c3 = st.columns([3, 2, 3])
    with c2:
        if os.path.exists("logo_taxbase.png"):
            st.image("logo_taxbase.png", use_container_width=True)
        else:
            st.markdown(f"<h1 style='text-align: center; color: {COR_PRIMARIA};'>TAXBASE</h1>", unsafe_allow_html=True)

    st.markdown(
        f"""<div style="text-align: center; margin-top: -15px;"><h4 style='color: {COR_SECUNDARIA}; font-weight: 300;'>Monitoramento de Atendimentos | Messenger</h4></div>""",
        unsafe_allow_html=True)

    # Controles de Admin e Logout
    c_blank, c_admin, c_logout = st.columns([8, 0.5, 0.5])
    if 'show_config' not in st.session_state: st.session_state['show_config'] = False

    with c_admin:
        if st.session_state['user_role'] == 'admin':
            if st.button("⚙️", help="Configurações"):
                st.session_state['show_config'] = not st.session_state['show_config']
                st.rerun()
    with c_logout:
        if st.button("🚪", help="Sair"):
            st.session_state['logged_in'] = False
            st.rerun()

    # --- LÓGICA DE EXIBIÇÃO ---

    if st.session_state.get('show_config', False) and st.session_state['user_role'] == 'admin':
        # TELA DE CONFIGURAÇÃO
        st.markdown("### ⚙️ Painel de Configurações")
        if st.button("⬅️ Voltar"):
            st.session_state['show_config'] = False
            st.rerun()

        if 'feedback_sucesso' in st.session_state:
            st.success(st.session_state['feedback_sucesso'])
            del st.session_state['feedback_sucesso']

        c1, c2 = st.columns(2)
        with c1:
            st.info("**Iniciar um Novo Mês**")
            nm = st.text_input("Nome (ex: 2026_03)", key="novo_mes_nome")
            up = st.file_uploader("CSV Inicial", key="novo_mes_arq")

            if st.button("Criar Mês"):
                if nm and up:
                    if not nm.endswith(".csv"): nm += ".csv"
                    with open(os.path.join("data", nm), "wb") as f:
                        f.write(up.getbuffer())

                    st.session_state['feedback_sucesso'] = f"Mês {nm} criado com sucesso!"
                    if "novo_mes_nome" in st.session_state: del st.session_state["novo_mes_nome"]
                    if "novo_mes_arq" in st.session_state: del st.session_state["novo_mes_arq"]
                    st.rerun()

        with c2:
            st.warning("**Atualizar Base Global**")
            up_xlsx = st.file_uploader("statusContatos.xlsx", type=['xlsx'], key="global_xlsx")
            if up_xlsx and st.button("Salvar Base"):
                with open("statusContatos.xlsx", "wb") as f: f.write(up_xlsx.getbuffer())
                st.success("Atualizado!")

    else:
        # TELA DE DASHBOARD
        dados_anos = listar_arquivos_por_ano()
        lista_anos = list(dados_anos.keys())

        abas_topo_nomes = lista_anos + ["📈 Análise Personalizada"]
        abas_topo = st.tabs(abas_topo_nomes)

        for i, ano in enumerate(lista_anos):
            with abas_topo[i]:
                meses_do_ano = dados_anos[ano]
                nomes_meses = [m['display'] for m in meses_do_ano]

                if not nomes_meses:
                    st.warning("Sem dados para este ano.")
                else:
                    abas_meses = st.tabs(nomes_meses)
                    for j, aba_mes_nome in enumerate(nomes_meses):
                        with abas_meses[j]:
                            caminho = meses_do_ano[j]['caminho']
                            df = carregar_dados_mes(caminho)

                            if df is not None:
                                # AQUI ESTÁ A MUDANÇA: Passamos o caminho do arquivo
                                renderizar_metricas_limpas(df, f"{ano}_{j}", caminho_arquivo_atual=caminho)

        # Aba de Análise Personalizada
        with abas_topo[-1]:
            st.markdown("### 🔍 Análise de Período Personalizado")
            arquivos = glob.glob(os.path.join("data", "*.csv"))

            all_dfs = []
            for f in arquivos:
                d = carregar_dados_mes(f)
                if d is not None: all_dfs.append(d)

            if not all_dfs:
                st.warning("Sem dados.")
            else:
                df_full = pd.concat(all_dfs).sort_values(by='Data')

                c_p1, c_p2 = st.columns(2)
                with c_p1:
                    periodo = st.selectbox("Selecione o Período",
                                           ["Últimos 3 Meses", "Últimos 6 Meses", "Ano Atual (2026)", "Todo o Período",
                                            "Personalizado..."])

                df_filtrado = df_full.copy()
                hoje = datetime.now()

                if periodo == "Últimos 3 Meses":
                    corte = hoje - timedelta(days=90)
                    df_filtrado = df_full[df_full['Data'] >= corte]
                elif periodo == "Últimos 6 Meses":
                    corte = hoje - timedelta(days=180)
                    df_filtrado = df_full[df_full['Data'] >= corte]
                elif periodo == "Ano Atual (2026)":
                    df_filtrado = df_full[df_full['Data'].dt.year == 2026]
                elif periodo == "Todo o Período":
                    df_filtrado = df_full
                elif periodo == "Personalizado...":
                    with c_p2:
                        min_d, max_d = df_full['Dia'].min(), df_full['Dia'].max()
                        dates = st.date_input("Início e Fim", [min_d, max_d])
                        if len(dates) == 2:
                            d_inicio = pd.to_datetime(dates[0])
                            d_fim = pd.to_datetime(dates[1]) + timedelta(days=1) - timedelta(seconds=1)
                            df_filtrado = df_full[(df_full['Data'] >= d_inicio) & (df_full['Data'] <= d_fim)]

                if df_filtrado.empty:
                    st.warning("Nenhum dado encontrado.")
                else:
                    st.info(
                        f"Analisando de **{df_filtrado['Dia'].min()}** até **{df_filtrado['Dia'].max()}** ({len(df_filtrado)} registros)")
                    renderizar_metricas_limpas(df_filtrado, "personalizado")


# --- EXECUÇÃO ---
if __name__ == "__main__":
    if autenticar():
        main_app()