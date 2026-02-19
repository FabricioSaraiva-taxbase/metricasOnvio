"""
Serviço de mapeamento Contato → Cliente (De/Para).
Portado de app_metricas.py (linhas 788-807 e 960-997).
"""

import os
import pandas as pd
from openpyxl import load_workbook
from typing import Tuple
from datetime import datetime

from backend.core.config import get_settings
from backend.services.bigquery_service import get_bq_client
from google.cloud import bigquery

TABLE_ID = "metricas.config_client_mapping"


def aplicar_depara(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica o cruzamento De/Para no DataFrame de atendimentos.

    1. Normaliza a coluna 'Contato' → 'Contato_Clean' (UPPER + STRIP)
    2. Lê o statusContatos.xlsx
    3. Faz merge para preencher 'Cliente_Final'
    4. Fallback → "NÃO IDENTIFICADO"
    """
    settings = get_settings()

    # 1. Normalização do Contato
    col_contato = "Contato" if "Contato" in df.columns else df.columns[1]
    df["Contato_Clean"] = df[col_contato].apply(
        lambda x: str(x).strip().upper() if pd.notnull(x) else ""
    )

    # 2. Cruzamento com Excel (Original) ou BigQuery (Novo)
    # Tenta carregar do BigQuery
    df_map = None
    cols_map = {}
    
    try:
        client = get_bq_client()
        if client:
            # Pega também updated_at para deduplicar
            query = f"SELECT original_name, normalized_name, updated_at FROM `{client.project}.{TABLE_ID}`"
            df_bq = client.query(query).to_dataframe()
            
            # --- DEDUPLICAÇÃO POR TIMESTAMP (LATEST WINS) ---
            if not df_bq.empty and "updated_at" in df_bq.columns:
                # Converte para datetime se não for
                df_bq["updated_at"] = pd.to_datetime(df_bq["updated_at"], errors="coerce")
                # Ordena decrescente (mais recente primeiro)
                df_bq = df_bq.sort_values(by="updated_at", ascending=False)
                # Mantém apenas a primeira ocorrência (a maiscente) de cada original_name
                df_bq = df_bq.drop_duplicates(subset=["original_name"], keep="first")
            
            # Rename to match expected logic below or standardise
            df_bq = df_bq.rename(columns={"original_name": "Nome_Map_Clean", "normalized_name": "Cliente_Alvo"})
            # Ensure upper/strip just in case (though migration did it)
            df_map_clean = df_bq.copy()
    except Exception as e:
        print(f"Error loading mapping from BQ: {e}")

    # Fallback to Excel if BQ failed
    if df_map_clean is None:
        map_path = settings.MAPPING_FILE
        if os.path.exists(map_path):
            try:
                df_map = pd.read_excel(map_path, engine="openpyxl")
                cols_map = {c.upper(): c for c in df_map.columns}

                if "NOME DO CONTATO" in cols_map and "NOME CLIENTE" in cols_map:
                    df_map_clean = df_map[
                        [cols_map["NOME DO CONTATO"], cols_map["NOME CLIENTE"]]
                    ].copy()
                    df_map_clean.columns = ["Nome_Map", "Cliente_Alvo"]
                    df_map_clean["Nome_Map_Clean"] = df_map_clean["Nome_Map"].apply(
                        lambda x: str(x).strip().upper() if pd.notnull(x) else ""
                    )
                    df_map_clean = df_map_clean.drop_duplicates(subset=["Nome_Map_Clean"])
            except Exception:
                pass
    
    if df_map_clean is not None:
         # Perform Merge
         try:
            df = pd.merge(
                df,
                df_map_clean,
                left_on="Contato_Clean",
                right_on="Nome_Map_Clean",
                how="left",
            )
            df["Cliente_Final"] = df["Cliente_Alvo"].fillna("NÃO IDENTIFICADO")
         except Exception:
            df["Cliente_Final"] = "NÃO IDENTIFICADO"
    else:
        df["Cliente_Final"] = "NÃO IDENTIFICADO"

    return df


def cadastrar_novo_cliente(nome_contato: str, nome_cliente: str) -> Tuple[bool, str]:
    """
    Cadastra um novo vínculo Contato → Cliente no statusContatos.xlsx.
    Usa o padrão APPEND-ONLY no BigQuery para evitar problemas de DML no Sandbox.
    """
    settings = get_settings()
    map_path = settings.MAPPING_FILE
    
    original = nome_contato.strip().upper()
    normalized = nome_cliente.strip()

    cadastrado = False
    
    # 1. Update Excel (Local Backup)
    if os.path.exists(map_path):
        try:
            wb = load_workbook(map_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            headers_upper = [str(h).upper() if h else "" for h in headers]
            try:
                idx_nome = headers_upper.index("NOME DO CONTATO")
                idx_cliente = headers_upper.index("NOME CLIENTE")
                
                nova_linha = [None] * len(headers)
                nova_linha[idx_nome] = original
                nova_linha[idx_cliente] = normalized

                ws.append(nova_linha)
                wb.save(map_path)
                cadastrado = True
            except ValueError:
                pass
        except Exception as e:
            pass 

    # 2. Update BigQuery (APPEND ONLY)
    try:
        client = get_bq_client()
        if client:
            print(f"[DEBUG] Tentando inserir no BQ: {original} -> {normalized}")
            
            # Não fazemos mais DELETE. Apenas INSERT da nova versão via LoadJob (Append).
            # O aplicar_depara vai pegar o registro com maior updated_at.
            
            row = {
                "original_name": original, 
                "normalized_name": normalized,
                "updated_at": datetime.utcnow().isoformat()
            }
            
            # Create DataFrame and convert to datetime to avoid pyarrow error
            df_insert = pd.DataFrame([row])
            df_insert["updated_at"] = pd.to_datetime(df_insert["updated_at"])
            
            print(f"[DEBUG] DataFrame para insert: {df_insert}")
            
            job_config = bigquery.LoadJobConfig(
                write_disposition="WRITE_APPEND",
            )
            job = client.load_table_from_dataframe(df_insert, f"{client.project}.{TABLE_ID}", job_config=job_config)
            job.result()
            
            print("[DEBUG] Sucesso no LoadJob BQ")
            return True, "Cadastrado no BigQuery (+ Excel Local) com sucesso!"
        else:
             print("[DEBUG] Client BQ nulo")

    except Exception as e:
        print(f"[DEBUG] BQ Insert Error: {e}")
        if not cadastrado:
            return False, f"Erro ao cadastrar: {e}"
            
    if cadastrado:
        return True, "Cadastrado no Excel Local (BQ indisponível)."
    
    return False, "Falha ao cadastrar."
