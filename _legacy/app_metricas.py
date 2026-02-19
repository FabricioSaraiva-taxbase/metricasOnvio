import streamlit as st
import json
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import os
import glob
import calendar
from datetime import datetime, timedelta
from openpyxl import load_workbook  # IMPORTANTE: Para salvar sem perder formatação

# --- INTEGRAÇÃO BIGQUERY (CONFIGURAÇÃO AUTOMÁTICA) ---
HAS_BQ = False
try:
    from google.cloud import bigquery
    from google.oauth2 import service_account
    
    HAS_BQ = True

    def get_bq_client():
        """Retorna cliente BQ autenticado via arquivo local ou Secrets."""
        try:
            # 1. Arquivo Local (Prioridade Dev)
            if os.path.exists("service_account.json"):
                os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "service_account.json"
                return bigquery.Client()
            
            # 2. Secrets do Streamlit Cloud (Prioridade Prod)
            if "gcp_service_account" in st.secrets:
                info = st.secrets["gcp_service_account"]
                creds = service_account.Credentials.from_service_account_info(info)
                return bigquery.Client(credentials=creds, project=creds.project_id)
            
            return None
        except Exception:
            return None

except ImportError:
    HAS_BQ = False
    def get_bq_client(): return None


# --- CONFIGURAÇÃO INICIAL ---
st.set_page_config(page_title="Métricas Taxbase", page_icon="🏢", layout="wide", initial_sidebar_state="collapsed")

# --- 0. GESTÃO DE USUÁRIOS (CREDENCIAIS) ---
CREDENCIAIS = {
    "fabricio": "admin",
    "fernando": "viewer",
    "gustavo": "viewer",
    "admin": "admin"
}
SENHA_PADRAO = "taxbase123"

# --- PALETA DE CORES TAXBASE (EXTRAÍDA DO LOGO) ---
TAXBASE_BLUE = "#00A0E3"      # Azul vibrante do logo
TAXBASE_SLATE = "#3C4A54"     # Cinza escuro do texto do logo
TAXBASE_LIGHT = "#F0F2F6"     # Cinza claro para fundos
TAXBASE_WHITE = "#FFFFFF"     # Branco

# TV MODE COLORS (Dark Theme)
TV_BG_PRIMARY = "#1E2328"     # Fundo principal escuro
TV_BG_SECONDARY = "#282D33"   # Fundo secundário (cards)
TV_TEXT = "#FFFFFF"           # Texto branco
TV_TEXT_MUTED = "#8B949E"     # Texto secundário

# --- CONSTANTES VISUAIS (MANTIDAS PARA COMPATIBILIDADE) ---
COR_PRIMARIA = TAXBASE_BLUE
COR_SECUNDARIA = TAXBASE_SLATE
COR_CINZA_CLARO = TAXBASE_LIGHT


def get_css_styles(tv_mode=False, dark_mode=False):
    """Retorna os estilos CSS baseados no modo (TV ou Desktop) e tema (Dark ou Light)."""
    
    # Dark mode é controlado independentemente - TV mode NÃO força dark mode
    # Usuário pode estar em TV mode com tema claro ou desktop mode com tema escuro
    use_dark = dark_mode
    
    css = ""
    
    if use_dark:
        # DARK MODE - Somente cores, sem alterar font-size
        css += f"""
        <style>
        /* === DARK MODE STYLES === */
        
        /* 1. GLOBAL DARK THEME */
        .stApp {{
            background-color: {TV_BG_PRIMARY} !important;
        }}
        .block-container {{
            background-color: {TV_BG_PRIMARY} !important;
            padding-top: 1rem;
            padding-bottom: 1rem;
            max-width: 100% !important;
        }}
        
        /* 2. HIDE SIDEBAR AND STREAMLIT CHROME */
        section[data-testid="stSidebar"] {{ display: none !important; }}
        header[data-testid="stHeader"] {{ display: none !important; }}
        .stDeployButton {{ display: none !important; }}
        #MainMenu {{ display: none !important; }}
        footer {{ display: none !important; }}
        
        /* 3. TYPOGRAPHY - COR APENAS, sem alterar font-size */
        h1, h2, h3 {{ color: {TV_TEXT} !important; }}
        h4 {{ color: {TAXBASE_BLUE} !important; font-weight: 600 !important; }}
        p, span, label, .stMarkdown {{ color: {TV_TEXT} !important; }}
        
        /* 4. KPI METRICS */
        div[data-testid="stMetric"] {{
            background: linear-gradient(135deg, {TV_BG_SECONDARY} 0%, {TV_BG_PRIMARY} 100%);
            border: 1px solid rgba(0,160,227,0.25);
            border-radius: 16px;
            padding: 1.25rem !important;
            box-shadow: 0 4px 20px rgba(0, 160, 227, 0.15);
        }}
        div[data-testid="stMetricValue"] {{
            color: {TAXBASE_BLUE} !important;
            font-weight: 700 !important;
        }}
        div[data-testid="stMetricLabel"] {{
            color: {TV_TEXT_MUTED} !important;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        /* 5. TABS STYLING */
        .stTabs [data-baseweb="tab-list"] {{
            background-color: {TV_BG_SECONDARY};
            border-radius: 12px;
            padding: 8px;
            gap: 4px;
            border-bottom: none !important;
        }}
        .stTabs [data-baseweb="tab"] {{
            background-color: transparent !important;
            color: {TV_TEXT_MUTED} !important;
            border: none !important;
            border-radius: 8px;
            padding: 12px 24px !important;
            font-size: 1.1rem !important;
            font-weight: 500;
        }}
        .stTabs [data-baseweb="tab"]:hover {{
            background-color: {TAXBASE_BLUE}20 !important;
            color: {TV_TEXT} !important;
        }}
        .stTabs [aria-selected="true"] {{
            background-color: {TAXBASE_BLUE} !important;
            color: {TV_TEXT} !important;
            font-weight: 700 !important;
        }}
        div[data-baseweb="tab-highlight"] {{ display: none !important; }}
        button[data-baseweb="tab"][aria-selected="true"] p {{
            color: {TV_TEXT} !important;
        }}
        
        /* 6. BUTTONS */
        .stButton > button {{
            background: linear-gradient(135deg, {TAXBASE_BLUE} 0%, #0080B3 100%) !important;
            color: white !important;
            border: none !important;
            border-radius: 10px !important;
            padding: 0.75rem 1.5rem !important;
            font-size: 1rem !important;
            font-weight: 600 !important;
            transition: all 0.3s ease !important;
            box-shadow: 0 4px 15px {TAXBASE_BLUE}40 !important;
        }}
        .stButton > button:hover {{
            transform: translateY(-2px) !important;
            box-shadow: 0 6px 20px {TAXBASE_BLUE}60 !important;
        }}
        
        /* 7. INPUTS - ROUNDED CORNERS */
        .stSelectbox, .stTextInput, .stDateInput {{
            background-color: transparent !important;
        }}
        .stSelectbox > div > div,
        .stTextInput > div > div {{
            background-color: {TV_BG_SECONDARY} !important;
            border-color: {TAXBASE_BLUE}40 !important;
            color: {TV_TEXT} !important;
            border-radius: 10px !important;
        }}
        /* Wildcard: TODOS os divs dentro de selectbox */
        .stSelectbox div {{
            border-radius: 10px !important;
        }}
        .stMultiSelect div {{
            border-radius: 10px !important;
        }}
        div[data-testid="stSelectbox"] div {{
            border-radius: 10px !important;
        }}
        div[data-baseweb="select"] > div {{
            border-radius: 10px !important;
        }}
        div[data-baseweb="select"] > div > div {{
            border-radius: 10px !important;
        }}
        div[data-baseweb="input"] {{
            border-radius: 10px !important;
        }}
        div[data-baseweb="input"] > div {{
            border-radius: 10px !important;
        }}
        div[data-baseweb="base-input"] {{
            border-radius: 10px !important;
        }}
        /* Dropdown popover e listbox */
        ul[role="listbox"] {{
            border-radius: 10px !important;
            background-color: {TV_BG_SECONDARY} !important;
        }}
        li[role="option"]:hover {{
            background-color: rgba(0,160,227,0.2) !important;
        }}
        div[data-baseweb="popover"] {{
            border-radius: 10px !important;
            overflow: hidden !important;
        }}
        div[data-baseweb="popover"] > div {{
            border-radius: 10px !important;
            background-color: {TV_BG_SECONDARY} !important;
        }}
        div[data-baseweb="menu"] {{
            border-radius: 10px !important;
        }}
        li[role="option"] {{
            border-radius: 6px !important;
        }}
        
        /* 8. TOOLTIPS - FIX LEGIBILITY */
        div[data-baseweb="tooltip"] {{
            background-color: {TAXBASE_WHITE} !important;
            color: {TAXBASE_SLATE} !important;
            border-radius: 8px !important;
        }}
        div[data-baseweb="tooltip"] * {{
            color: {TAXBASE_SLATE} !important;
        }}
        
        /* 9. FILE UPLOADER - FIX TEXT */
        .stFileUploader {{
            background-color: {TV_BG_SECONDARY} !important;
            border-radius: 10px !important;
        }}
        .stFileUploader label,
        .stFileUploader span,
        .stFileUploader p,
        .stFileUploader small {{
            color: {TV_TEXT} !important;
        }}
        section[data-testid="stFileUploader"] {{
            background-color: {TV_BG_SECONDARY} !important;
            border-radius: 10px !important;
            padding: 1rem !important;
        }}
        section[data-testid="stFileUploader"] * {{
            color: {TV_TEXT} !important;
        }}
        
        /* 8. EXPANDERS */
        .streamlit-expanderHeader {{
            background-color: {TV_BG_SECONDARY} !important;
            color: {TV_TEXT} !important;
            border-radius: 10px !important;
            font-size: 1.1rem !important;
        }}
        .streamlit-expanderContent {{
            background-color: {TV_BG_SECONDARY} !important;
            border: 1px solid {TAXBASE_BLUE}30 !important;
            border-radius: 0 0 10px 10px !important;
        }}
        
        /* 9. DATAFRAME */
        .stDataFrame {{
            background-color: {TV_BG_SECONDARY} !important;
            border-radius: 12px !important;
        }}
        
        /* 10. DIVIDERS */
        hr {{
            border-color: {TAXBASE_BLUE}30 !important;
        }}
        
        /* 11. CHARTS BACKGROUND */
        .js-plotly-plot {{
            background-color: transparent !important;
        }}
        
        /* 12. SCROLLBAR STYLING */
        ::-webkit-scrollbar {{
            width: 8px;
            height: 8px;
        }}
        ::-webkit-scrollbar-track {{
            background: {TV_BG_SECONDARY};
        }}
        ::-webkit-scrollbar-thumb {{
            background: {TAXBASE_BLUE};
            border-radius: 4px;
        }}
        
        </style>
        """
    else:
        # LIGHT MODE - Light theme, standard fonts
        css += f"""
        <style>
        /* === LIGHT MODE STYLES === */
        
        /* 1. LAYOUT GERAL */
        .block-container {{ 
            padding-top: 1rem; 
            padding-bottom: 1rem;
            max-width: 100% !important;
        }}
        section[data-testid="stSidebar"] {{ display: none; }}
        header[data-testid="stHeader"] {{ display: none !important; }}
        .stDeployButton {{ display: none !important; }}
        #MainMenu {{ display: none !important; }}
        footer {{ display: none !important; }}
        
        /* 2. TYPOGRAPHY */
        h1, h2, h3 {{ 
            color: {TAXBASE_SLATE}; 
            font-family: 'Segoe UI', -apple-system, sans-serif; 
        }}
        h4 {{ 
            color: {TAXBASE_BLUE}; 
            font-weight: 600;
        }}
        
        /* 3. KPI METRICS */
        div[data-testid="stMetric"] {{
            background: linear-gradient(135deg, {TAXBASE_WHITE} 0%, {TAXBASE_LIGHT} 100%);
            border: 1px solid {TAXBASE_BLUE}30;
            border-radius: 12px;
            padding: 1.25rem !important;
            box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
            transition: all 0.3s ease;
        }}
        div[data-testid="stMetric"]:hover {{
            transform: translateY(-2px);
            box-shadow: 0 4px 20px rgba(0, 160, 227, 0.2);
        }}
        div[data-testid="stMetricValue"] {{ 
            color: {TAXBASE_BLUE}; 
            font-weight: 700;
            font-size: 2rem;
        }}
        div[data-testid="stMetricLabel"] {{
            color: {TAXBASE_SLATE};
            font-weight: 500;
            text-transform: uppercase;
            font-size: 0.85rem;
            letter-spacing: 0.5px;
        }}
        
        /* 4. TABS - CLEAN STYLE */
        .stTabs [data-baseweb="tab-list"] {{
            background-color: {TAXBASE_LIGHT};
            border-radius: 10px;
            padding: 6px;
            gap: 4px;
            border-bottom: none !important;
        }}
        .stTabs [data-baseweb="tab"] {{
            background-color: transparent !important;
            border: none !important;
            border-radius: 8px;
            color: {TAXBASE_SLATE};
            padding: 10px 20px !important;
            font-weight: 500;
            transition: all 0.2s ease;
        }}
        .stTabs [data-baseweb="tab"]:hover {{
            background-color: {TAXBASE_BLUE}15 !important;
        }}
        .stTabs [aria-selected="true"] {{
            background-color: {TAXBASE_WHITE} !important;
            color: {TAXBASE_BLUE} !important;
            font-weight: 700 !important;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
        }}
        div[data-baseweb="tab-highlight"] {{ display: none !important; }}
        button[data-baseweb="tab"][aria-selected="true"] p {{
            color: {TAXBASE_BLUE} !important;
        }}
        
        /* 5. BUTTONS */
        .stButton > button {{
            background: linear-gradient(135deg, {TAXBASE_BLUE} 0%, #0080B3 100%);
            color: white;
            border: none;
            border-radius: 8px;
            padding: 0.5rem 1.25rem;
            font-weight: 600;
            transition: all 0.3s ease;
            box-shadow: 0 2px 8px {TAXBASE_BLUE}30;
        }}
        .stButton > button:hover {{
            transform: translateY(-1px);
            box-shadow: 0 4px 12px {TAXBASE_BLUE}50;
        }}
        
        /* 6. INPUTS */
        div[data-baseweb="select"]:focus-within > div, 
        div[data-baseweb="input"]:focus-within > div {{
            border-color: {TAXBASE_BLUE} !important;
            box-shadow: 0 0 0 1px {TAXBASE_BLUE} !important;
        }}
        input {{ caret-color: {TAXBASE_BLUE} !important; }}
        div[data-baseweb="checkbox"] span[aria-checked="true"] {{
            background-color: {TAXBASE_BLUE} !important;
            border-color: {TAXBASE_BLUE} !important;
        }}
        
        /* 7. EXPANDERS */
        .streamlit-expanderHeader {{
            background-color: {TAXBASE_LIGHT};
            border-radius: 8px;
            font-weight: 600;
            color: {TAXBASE_SLATE};
        }}
        
        /* 8. DIVIDERS */
        hr {{
            border-color: {TAXBASE_LIGHT} !important;
            margin: 1rem 0 !important;
        }}
        
        </style>
        """
    
    # ============================================================
    #  BLOCO 2 – ESCALA TV (fontes maiores, aplicado POR CIMA)
    # ============================================================
    if tv_mode:
        css += f"""
        <style>
        /* === TV SCALE – FONTES MAIORES === */
        h1 {{ font-size: 3rem !important; font-weight: 700 !important; }}
        h2 {{ font-size: 2.2rem !important; }}
        h3 {{ font-size: 1.8rem !important; }}
        h4 {{ font-size: 1.5rem !important; }}
        p, span, label, .stMarkdown {{ font-size: 1.1rem !important; }}
        
        div[data-testid="stMetricValue"] {{
            font-size: 3.5rem !important;
            font-weight: 800 !important;
        }}
        div[data-testid="stMetricLabel"] {{
            font-size: 1.2rem !important;
            letter-spacing: 1px;
        }}
        div[data-testid="stMetricDelta"] {{
            font-size: 1.1rem !important;
        }}
        
        .stTabs [data-baseweb="tab"] {{
            padding: 12px 24px !important;
            font-size: 1.1rem !important;
        }}
        
        .stButton > button {{
            padding: 0.75rem 1.5rem !important;
            font-size: 1rem !important;
        }}
        
        .streamlit-expanderHeader {{
            font-size: 1.1rem !important;
        }}
        </style>
        """
    
    return css


def render_header(tv_mode=False):
    """Renderiza o header com logo, título e controles."""
    
    # Container do header
    col_logo, col_title, col_controls = st.columns([1, 3, 1])
    
    with col_logo:
        if os.path.exists("logo_taxbase.png"):
            st.image("logo_taxbase.png", width=150 if not tv_mode else 200)
    
    with col_title:
        if tv_mode:
            st.markdown(f"""
                <div style="text-align: center; padding: 10px 0;">
                    <h2 style="margin: 0; color: {TV_TEXT}; font-weight: 300;">
                        Monitoramento de Atendimentos
                    </h2>
                </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
                <div style="text-align: center; padding: 5px 0;">
                    <p style="margin: 0; color: {TAXBASE_SLATE}; font-size: 1.1rem; font-weight: 400;">
                        Monitoramento de Atendimentos | Messenger
                    </p>
                </div>
            """, unsafe_allow_html=True)
    
    with col_controls:
        c1, c2, c3, c4 = st.columns(4)
        
        with c1:
            # TV Mode Toggle
            tv_icon = "🖥️" if st.session_state.get('tv_mode', False) else "📺"
            tv_label = "Desktop" if st.session_state.get('tv_mode', False) else "TV"
            if st.button(f"{tv_icon}", help=f"Alternar para modo {tv_label}", key="btn_tv_toggle"):
                st.session_state['tv_mode'] = not st.session_state.get('tv_mode', False)
                st.rerun()
        
        with c2:
            # Dark Mode Toggle (independent)
            dark_icon = "☀️" if st.session_state.get('dark_mode', False) else "🌙"
            dark_label = "Modo Claro" if st.session_state.get('dark_mode', False) else "Modo Escuro"
            if st.button(f"{dark_icon}", help=f"Alternar para {dark_label}", key="btn_dark_toggle"):
                st.session_state['dark_mode'] = not st.session_state.get('dark_mode', False)
                st.rerun()
        
        with c3:
            # Settings (Admin only)
            if st.session_state.get('user_role') == 'admin':
                if st.button("⚙️", help="Configurações", key="btn_config_header"):
                    st.session_state['show_config'] = not st.session_state.get('show_config', False)
                    st.rerun()
        
        with c4:
            # Logout
            if st.button("🚪", help="Sair", key="btn_logout_header"):
                st.session_state['logged_in'] = False
                st.rerun()


def render_empty_state(message, icon="📊", suggestion=None, tv_mode=False):
    """Renderiza um estado vazio estilizado."""
    
    bg_color = TV_BG_SECONDARY if tv_mode else TAXBASE_LIGHT
    text_color = TV_TEXT if tv_mode else TAXBASE_SLATE
    muted_color = TV_TEXT_MUTED if tv_mode else "#6B7280"
    
    st.markdown(f"""
        <div style="
            text-align: center;
            padding: 4rem 2rem;
            background: {bg_color};
            border-radius: 16px;
            margin: 2rem 0;
        ">
            <div style="font-size: 4rem; margin-bottom: 1rem;">{icon}</div>
            <h3 style="color: {text_color}; margin-bottom: 0.5rem;">{message}</h3>
            {f'<p style="color: {muted_color}; font-size: 1rem;">{suggestion}</p>' if suggestion else ''}
        </div>
    """, unsafe_allow_html=True)


def render_loading_state(message="Carregando dados...", tv_mode=False):
    """Renderiza um estado de carregamento estilizado."""
    
    text_color = TV_TEXT if tv_mode else TAXBASE_SLATE
    
    st.markdown(f"""
        <div style="
            text-align: center;
            padding: 2rem;
        ">
            <div style="
                width: 50px;
                height: 50px;
                border: 4px solid {TAXBASE_BLUE}30;
                border-top-color: {TAXBASE_BLUE};
                border-radius: 50%;
                animation: spin 1s linear infinite;
                margin: 0 auto 1rem;
            "></div>
            <p style="color: {text_color}; font-size: 1.1rem;">{message}</p>
        </div>
        <style>
            @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
        </style>
    """, unsafe_allow_html=True)


# --- 1. FUNÇÕES DE BACKEND ---

def autenticar():
    """Gerencia a tela de login com visual premium."""
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False
        st.session_state['user_role'] = None
        st.session_state['tv_mode'] = False
        st.session_state['dark_mode'] = False
        st.session_state['show_config'] = False
    
    if not st.session_state['logged_in']:
        # CSS premium para login
        st.markdown(f"""
            <style>
            /* Background gradiente animado */
            .stApp {{
                background: linear-gradient(135deg, #0f1923 0%, #1a2a3a 30%, #0d2137 60%, #152535 100%) !important;
                overflow: hidden;
            }}
            
            /* Esconder header/footer do Streamlit */
            header[data-testid="stHeader"] {{ display: none !important; }}
            .stDeployButton {{ display: none !important; }}
            #MainMenu {{ display: none !important; }}
            footer {{ display: none !important; }}
            
            /* Esconder sidebar */
            section[data-testid="stSidebar"] {{ display: none !important; }}
            
            /* Container principal */
            .main .block-container {{
                padding-top: 0 !important;
                padding-bottom: 0 !important;
                max-width: 100% !important;
            }}
            
            /* Inputs do form */
            .stTextInput > div > div {{
                background-color: rgba(255,255,255,0.92) !important;
                border: 1px solid rgba(0,160,227,0.3) !important;
                border-radius: 12px !important;
                transition: all 0.3s ease !important;
            }}
            .stTextInput > div > div:focus-within {{
                border-color: {TAXBASE_BLUE} !important;
                box-shadow: 0 0 0 2px rgba(0,160,227,0.2) !important;
                background-color: #FFFFFF !important;
            }}
            .stTextInput input {{
                color: #1E2328 !important;
            }}
            .stTextInput input::placeholder {{
                color: #8B949E !important;
            }}
            .stTextInput label {{
                color: rgba(255,255,255,0.8) !important;
                font-weight: 500 !important;
            }}
            
            /* Botão entrar */
            .stFormSubmitButton > button {{
                background: linear-gradient(135deg, {TAXBASE_BLUE} 0%, #0080B3 100%) !important;
                color: white !important;
                border: none !important;
                border-radius: 12px !important;
                padding: 0.85rem 2rem !important;
                font-size: 1.1rem !important;
                font-weight: 700 !important;
                letter-spacing: 0.5px !important;
                transition: all 0.3s ease !important;
                box-shadow: 0 4px 20px rgba(0,160,227,0.35) !important;
            }}
            .stFormSubmitButton > button:hover {{
                transform: translateY(-2px) !important;
                box-shadow: 0 8px 30px rgba(0,160,227,0.5) !important;
            }}
            
            /* Alertas */
            .stAlert {{
                border-radius: 12px !important;
            }}
            </style>
        """, unsafe_allow_html=True)
        
        # Layout centralizado
        col_spacer_l, col_center, col_spacer_r = st.columns([1.2, 1, 1.2])
        
        with col_center:
            # Espaçamento superior
            st.markdown("<div style='height: 12vh;'></div>", unsafe_allow_html=True)
            
            # Logo
            if os.path.exists("logo_taxbase.png"):
                st.image("logo_taxbase.png", width=180)
            else:
                st.markdown(f"""
                    <h1 style="
                        color: {TAXBASE_BLUE};
                        font-size: 2.2rem;
                        font-weight: 800;
                        margin: 0 0 0.25rem;
                        letter-spacing: 2px;
                    ">TAXBASE</h1>
                """, unsafe_allow_html=True)
            
            st.markdown(f"""
                <p style="
                    color: rgba(255,255,255,0.5);
                    margin: 0.5rem 0 2rem;
                    font-size: 0.95rem;
                    letter-spacing: 0.5px;
                ">
                    Painel de Métricas
                </p>
            """, unsafe_allow_html=True)
            
            # Form
            with st.form("login_form", clear_on_submit=False):
                usuario = st.text_input("👤 Usuário", placeholder="Digite seu usuário").lower().strip()
                senha = st.text_input("🔒 Senha", type="password", placeholder="Digite sua senha")
                
                st.markdown("<div style='height: 0.75rem;'></div>", unsafe_allow_html=True)
                
                submit = st.form_submit_button("Entrar →", use_container_width=True)
            
            if submit:
                if usuario and senha:
                    with st.spinner("Autenticando..."):
                        if usuario in CREDENCIAIS and senha == SENHA_PADRAO:
                            st.session_state['logged_in'] = True
                            st.session_state['user_role'] = CREDENCIAIS[usuario]
                            st.success("✅ Login realizado com sucesso!")
                            st.rerun()
                        else:
                            st.error("❌ Credenciais inválidas. Tente novamente.")
                else:
                    st.warning("⚠️ Preencha todos os campos.")
            
            # Rodapé
            st.markdown(f"""
                <p style="
                    text-align: center;
                    color: rgba(255,255,255,0.25);
                    font-size: 0.8rem;
                    margin-top: 2rem;
                ">
                    © 2025 Taxbase · Todos os direitos reservados
                </p>
            """, unsafe_allow_html=True)
        
        return False
    return True


@st.cache_data(ttl=3600, show_spinner=False)
def carregar_dados_mes(identificador):
    """
    HÍBRIDO COM CACHE: Lê do BigQuery ou CSV e guarda na memória para performance.
    """
    df = None

    # --- TENTATIVA 1: BIGQUERY (NUVEM) ---
    if isinstance(identificador, str) and identificador.startswith("BQ:") and HAS_BQ:
        try:
            partes = identificador.replace("BQ:", "").split('_')
            if len(partes) >= 2:
                ano, mes = partes[0], partes[1]
                tabela = f"atendimentos_{ano}_{mes}"
                client = get_bq_client()
                query = f"SELECT * FROM `taxbase-metricasmessenger.metricas.{tabela}`"
                df = client.query(query).to_dataframe()
        except Exception:
            pass

    # --- TENTATIVA 2: CSV LOCAL (DISCO) ---
    if df is None:
        if not os.path.exists(identificador): return None
        try:
            try:
                df = pd.read_csv(identificador, sep=';', encoding='utf-8')
            except:
                df = pd.read_csv(identificador, sep=';', encoding='latin1')
        except:
            return None

    if df is None: return None

    # --- HIGIENIZAÇÃO E REGRAS DE NEGÓCIO ---

    # 1. Contato
    col_contato = 'Contato' if 'Contato' in df.columns else df.columns[1]
    df['Contato_Clean'] = df[col_contato].apply(lambda x: str(x).strip().upper() if pd.notnull(x) else "")

    # 2. Cruzamento com Excel
    map_path = 'statusContatos.xlsx'
    if os.path.exists(map_path):
        try:
            df_map = pd.read_excel(map_path, engine='openpyxl')
            cols_map = {c.upper(): c for c in df_map.columns}
            if 'NOME DO CONTATO' in cols_map and 'NOME CLIENTE' in cols_map:
                df_map_clean = df_map[[cols_map['NOME DO CONTATO'], cols_map['NOME CLIENTE']]].copy()
                df_map_clean.columns = ['Nome_Map', 'Cliente_Alvo']
                df_map_clean['Nome_Map_Clean'] = df_map_clean['Nome_Map'].apply(
                    lambda x: str(x).strip().upper() if pd.notnull(x) else "")
                df_map_clean = df_map_clean.drop_duplicates(subset=['Nome_Map_Clean'])
                df = pd.merge(df, df_map_clean, left_on='Contato_Clean', right_on='Nome_Map_Clean', how='left')
                df['Cliente_Final'] = df['Cliente_Alvo'].fillna("NÃO IDENTIFICADO")
            else:
                df['Cliente_Final'] = "NÃO IDENTIFICADO"
        except:
            df['Cliente_Final'] = "NÃO IDENTIFICADO"
    else:
        df['Cliente_Final'] = "NÃO IDENTIFICADO"

    # 3. Data (Correção de Fuso e Formato)
    if 'Data' in df.columns:
        df['Data'] = pd.to_datetime(df['Data'], dayfirst=True, errors='coerce')
        if df['Data'].dt.tz is not None:
            df['Data'] = df['Data'].dt.tz_localize(None)
        df['Dia'] = df['Data'].dt.date
        df = df.sort_values(by='Data', ascending=False)

    return df


def listar_arquivos_por_ano():
    """
    HÍBRIDO INTELIGENTE: Lista Nuvem e Local, removendo duplicatas locais.
    """
    estrutura = {}
    meses_na_nuvem = set()  # Rastreia o que já baixamos da nuvem (Ex: "2026_01")

    if not os.path.exists("data"): os.makedirs("data")

    # 1. LISTAR DO BIGQUERY (NUVEM)
    if HAS_BQ:
        try:
            client = get_bq_client()
            dataset_ref = client.dataset("metricas")
            tables = list(client.list_tables(dataset_ref))

            for table in tables:
                tid = table.table_id
                if tid.startswith("atendimentos_"):
                    partes = tid.split('_')  # [atendimentos, 2026, 01]
                    if len(partes) >= 3:
                        ano, mes = partes[1], partes[2]
                        if ano not in estrutura: estrutura[ano] = []

                        # Marca que este mês já existe na nuvem
                        meses_na_nuvem.add(f"{ano}_{mes}")

                        estrutura[ano].append({
                            'caminho': f"BQ:{ano}_{mes}",
                            'display': f"{mes}/{ano[2:]} ☁️",
                            'mes_raw': int(mes) if mes.isdigit() else 0
                        })
        except:
            pass

    # 2. LISTAR DO DISCO (LOCAL) - COM FILTRO DE DUPLICIDADE
    arquivos = glob.glob(os.path.join("data", "*.csv"))
    for f in arquivos:
        nome_base = os.path.basename(f).replace(".csv", "")
        partes = nome_base.split('_', 1)  # Ex: 2026_01

        if len(partes) == 2:
            ano = partes[0]
            resto = partes[1]  # Ex: "01" ou "01 (PARCIAL)"
            mes_limpo = resto.split(' ')[0]  # Pega só o "01"

            # SE JÁ TEM NA NUVEM, IGNORA O LOCAL (EVITA DUPLICIDADE)
            if f"{ano}_{mes_limpo}" in meses_na_nuvem:
                continue

            if ano.isdigit() and len(ano) == 4:
                if ano not in estrutura: estrutura[ano] = []

                mes_num_str = resto[:2]
                try:
                    mes_int = int(mes_num_str)
                except:
                    mes_int = 0

                sufixo = resto[2:]
                display = f"{mes_num_str}/{ano[2:]}{sufixo}"  # Exibe normal (sem tag Nuvem)

                estrutura[ano].append({'caminho': f, 'display': display, 'mes_raw': mes_int})

    # Ordenação
    for ano in estrutura:
        estrutura[ano] = sorted(estrutura[ano], key=lambda x: x['mes_raw'], reverse=True)

    return dict(sorted(estrutura.items(), reverse=True))


# --- LABELS DE MESES (HÍBRIDO: BIGQUERY COM FALLBACK JSON LOCAL) ---

LABELS_JSON_PATH = "month_labels.json"

def _load_labels_json():
    """Carrega labels do arquivo JSON local."""
    if os.path.exists(LABELS_JSON_PATH):
        try:
            with open(LABELS_JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {}

def _save_labels_json(labels):
    """Salva labels no arquivo JSON local."""
    with open(LABELS_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(labels, f, ensure_ascii=False, indent=2)

def load_month_labels():
    """Carrega labels dos meses. Mescla BigQuery + JSON local (BQ tem prioridade)."""
    # Base: JSON local
    labels = _load_labels_json()
    
    # Tenta complementar/sobrescrever com BigQuery
    if HAS_BQ:
        try:
            client = get_bq_client()
            query = "SELECT mes_key, label FROM `taxbase-metricasmessenger.metricas.month_labels` WHERE label IS NOT NULL AND label != ''"
            df = client.query(query).to_dataframe()
            bq_labels = dict(zip(df['mes_key'], df['label']))
            labels.update(bq_labels)  # BQ sobrescreve JSON para mesmas chaves
        except:
            pass
    
    return labels

def save_month_label(mes_key, label):
    """Salva label. Tenta BigQuery, se falhar (billing) usa JSON local."""
    # Tenta BigQuery primeiro
    if HAS_BQ:
        try:
            client = get_bq_client()
            del_query = f"DELETE FROM `taxbase-metricasmessenger.metricas.month_labels` WHERE mes_key = '{mes_key}'"
            client.query(del_query).result()
            if label.strip():
                label_safe = label.replace("'", "\\'")
                ins_query = f"INSERT INTO `taxbase-metricasmessenger.metricas.month_labels` (mes_key, label) VALUES ('{mes_key}', '{label_safe}')"
                client.query(ins_query).result()
            return True, "Salvo no BigQuery"
        except Exception as e:
            # Se for erro de billing/DML, faz fallback para JSON
            if "billingNotEnabled" in str(e) or "free tier" in str(e).lower():
                pass  # Cai no fallback abaixo
            else:
                return False, str(e)
    
    # Fallback: JSON local
    try:
        labels = _load_labels_json()
        if label.strip():
            labels[mes_key] = label
        else:
            labels.pop(mes_key, None)
        _save_labels_json(labels)
        return True, "Salvo localmente (JSON)"
    except Exception as e:
        return False, str(e)

def cadastrar_novo_cliente(nome_contato, nome_cliente):
    """
    Função Feature 3 (CORRIGIDA):
    Usa openpyxl para fazer APPEND na planilha sem sobrescrever formatação.
    """
    map_path = 'statusContatos.xlsx'
    if not os.path.exists(map_path):
        return False, "Arquivo statusContatos.xlsx não encontrado."

    try:
        # Carrega o workbook existente (mantendo estilos)
        wb = load_workbook(map_path)
        ws = wb.active  # Pega a aba ativa

        # Descobre qual coluna é qual
        headers = [cell.value for cell in ws[1]]
        headers_upper = [str(h).upper() if h else "" for h in headers]

        try:
            idx_nome = headers_upper.index('NOME DO CONTATO')
            idx_cliente = headers_upper.index('NOME CLIENTE')
        except ValueError:
            return False, "Colunas 'Nome do Contato' ou 'NOME CLIENTE' não encontradas no cabeçalho."

        # Prepara a nova linha (lista de valores vazios do tamanho do header)
        nova_linha = [None] * len(headers)
        nova_linha[idx_nome] = nome_contato
        nova_linha[idx_cliente] = nome_cliente

        # Adiciona a linha ao final
        ws.append(nova_linha)

        # Salva o arquivo
        wb.save(map_path)
        return True, "Cadastrado com sucesso (Formatação preservada)!"

    except Exception as e:
        return False, str(e)


# --- 2. FRONTEND PRINCIPAL ---

def renderizar_metricas_limpas(df, titulo_contexto, caminho_arquivo_atual=None):
    """Componente reutilizável de dashboard com visual limpo e chaves únicas."""
    
    tv_mode = st.session_state.get('tv_mode', False)
    dark_mode = st.session_state.get('dark_mode', False)
    use_dark = dark_mode  # Dark mode é controlado independentemente

    # --- ÁREA DE CADASTRO RÁPIDO (MANTIDA) ---
    pendentes = df[df['Cliente_Final'] == "NÃO IDENTIFICADO"]
    if not pendentes.empty:
        with st.expander(f"🚨 Há {len(pendentes)} atendimentos 'NÃO IDENTIFICADO'. Clique para vincular agora.",
                         expanded=False):
            c_cad1, c_cad2, c_cad3 = st.columns([2, 2, 1])
            with c_cad1:
                lista_nomes = sorted(pendentes['Contato'].unique().astype(str))
                contato_selecionado = st.selectbox("Selecione o Contato", lista_nomes,
                                                   key=f"sel_cont_{titulo_contexto}")
            with c_cad2:
                empresa_input = st.text_input("Nome da Empresa (Cliente)", key=f"inp_cli_{titulo_contexto}",
                                              placeholder="Ex: PADARIA DO JOÃO")
            with c_cad3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("💾 Vincular", key=f"btn_vinc_{titulo_contexto}"):
                    if empresa_input:
                        sucesso, msg = cadastrar_novo_cliente(contato_selecionado, empresa_input.upper())
                        if sucesso:
                            st.success(msg); st.rerun()
                        else:
                            st.error(msg)
                    else:
                        st.warning("Digite o nome da empresa.")

    st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)

    # --- FILTROS ---
    st.markdown("#### 🔍 Filtros" if not tv_mode else "", unsafe_allow_html=True)
    c_f1, c_f2 = st.columns([2, 2])
    with c_f1:
        analistas = ["Todos"] + sorted(list(df['Atendido por'].unique())) if 'Atendido por' in df.columns else ["Todos"]
        filtro_analista = st.selectbox(f"Analista", analistas, key=f"sel_an_{titulo_contexto}")

    with c_f2:
        clientes_validos = sorted(
            [c for c in df['Cliente_Final'].unique() if c not in ["TAXBASE INTERNO", "IGNORAR", "NÃO IDENTIFICADO"]])
        clientes_opts = ["Todos (Visão Geral)"] + clientes_validos
        filtro_cliente = st.selectbox("Empresa", clientes_opts, key=f"sel_cli_{titulo_contexto}")

    # APLICAÇÃO DOS FILTROS
    df_view = df.copy()
    if filtro_analista != "Todos": df_view = df_view[df_view['Atendido por'] == filtro_analista]
    if filtro_cliente != "Todos (Visão Geral)": df_view = df_view[df_view['Cliente_Final'] == filtro_cliente]

    df_ranking = df_view[~df_view['Cliente_Final'].isin(["TAXBASE INTERNO", "IGNORAR", "NÃO IDENTIFICADO"])]

    # --- KPIS ---
    st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
    k1, k2, k3 = st.columns(3)
    
    total_bruto = len(df_view)
    total_validos = len(df_ranking)
    clientes_unicos = df_ranking['Cliente_Final'].nunique()
    
    k1.metric("📊 Total Geral", f"{total_bruto:,}".replace(",", "."))
    k2.metric("✅ Atendimentos Válidos", f"{total_validos:,}".replace(",", "."))
    k3.metric("🏢 Clientes Únicos", f"{clientes_unicos:,}".replace(",", "."))
    
    st.markdown("<hr style='margin: 1rem 0;'>", unsafe_allow_html=True)

    # --- GRÁFICOS ---
    c_g1, c_g2 = st.columns([2, 1])
    
    with c_g1:
        if filtro_cliente != "Todos (Visão Geral)":
            st.markdown(f"#### 📈 Volume: {filtro_cliente}")
        else:
            st.markdown("#### 🏆 Clientes com Maior Demanda")

        if not df_ranking.empty:
            ranking = df_ranking['Cliente_Final'].value_counts().head(10).reset_index()
            ranking.columns = ['Cliente', 'Volume']
            
            # Chart styling based on mode
            if use_dark:
                fig_bar = px.bar(
                    ranking, 
                    x='Volume', 
                    y='Cliente', 
                    orientation='h', 
                    text='Volume',
                    color_discrete_sequence=[TAXBASE_BLUE]
                )
                fig_bar.update_layout(
                    yaxis={'categoryorder': 'total ascending', 'title': ''},
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=0, r=0, t=10, b=0),
                    font=dict(color=TV_TEXT, size=14),
                    xaxis=dict(gridcolor='rgba(0, 160, 227, 0.2)', title=''),
                    showlegend=False,
                    height=400
                )

                fig_bar.update_traces(
                    textposition='outside',
                    textfont=dict(size=16, color=TV_TEXT)
                )
            else:
                fig_bar = px.bar(
                    ranking, 
                    x='Volume', 
                    y='Cliente', 
                    orientation='h', 
                    text='Volume',
                    color_discrete_sequence=[TAXBASE_BLUE]
                )
                fig_bar.update_layout(
                    yaxis={'categoryorder': 'total ascending', 'title': ''}, 
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=0, r=0, t=10, b=0),
                    xaxis=dict(gridcolor='#E5E7EB', title=''),
                    showlegend=False,
                    height=350
                )

                fig_bar.update_traces(textposition='outside')

            st.plotly_chart(fig_bar, use_container_width=True, key=f"bar_{titulo_contexto}")

            # RANKING COMPLETO (PRESERVADO)
            if filtro_cliente == "Todos (Visão Geral)":
                with st.expander("📋 Ver Ranking Completo de Clientes"):
                    ranking_full = df_ranking['Cliente_Final'].value_counts().reset_index()
                    ranking_full.columns = ['Cliente', 'Volume']
                    st.dataframe(ranking_full, use_container_width=True, hide_index=True)
        else:
            render_empty_state(
                "Nenhum dado encontrado",
                icon="📭",
                suggestion="Ajuste os filtros para visualizar os dados",
                tv_mode=use_dark
            )

    with c_g2:
        st.markdown("#### 📅 Atendimentos por Dia")
        
        # CHECKBOX DO CALENDÁRIO COMPLETO (PRESERVADO)
        usar_mes_completo = st.checkbox(
            "📆 Exibir calendário completo", 
            value=True, 
            key=f"chk_{titulo_contexto}",
            help="Inclui dias com zero atendimentos."
        )

        if not df_ranking.empty:
            min_date = df_ranking['Dia'].min()
            max_date = df_ranking['Dia'].max()

            if usar_mes_completo:
                start_date = min_date.replace(day=1)
                last_day_num = calendar.monthrange(max_date.year, max_date.month)[1]
                end_date = max_date.replace(day=last_day_num)
                full_range = pd.date_range(start=start_date, end=end_date)
                timeline = df_ranking.groupby('Dia').size()
                timeline.index = pd.to_datetime(timeline.index)
                timeline = timeline.reindex(full_range, fill_value=0).reset_index()
                timeline.columns = ['Dia', 'Volume']
                timeline['Dia'] = timeline['Dia'].dt.date
            else:
                timeline = df_ranking.groupby('Dia').size().reset_index(name='Volume')

            if use_dark:
                fig_line = px.line(
                    timeline, 
                    x='Dia', 
                    y='Volume', 
                    markers=True
                )
                fig_line.update_traces(
                    line_color=TAXBASE_BLUE,
                    line_width=3,
                    marker=dict(size=8, color=TAXBASE_BLUE)
                )
                fig_line.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=0, r=0, t=10, b=0),
                    font=dict(color=TV_TEXT, size=12),
                    xaxis=dict(gridcolor='rgba(0, 160, 227, 0.2)', title=''),
                    yaxis=dict(gridcolor='rgba(0, 160, 227, 0.2)', title=''),
                    showlegend=False,
                    height=300
                )
            else:
                fig_line = px.line(
                    timeline, 
                    x='Dia', 
                    y='Volume', 
                    markers=True
                )
                fig_line.update_traces(line_color=TAXBASE_BLUE)
                fig_line.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=0, r=0, t=10, b=0),
                    xaxis=dict(gridcolor='#E5E7EB', title=''),
                    yaxis=dict(gridcolor='#E5E7EB', title=''),
                    showlegend=False,
                    height=280
                )
            
            fig_line.update_yaxes(rangemode="tozero")
            st.plotly_chart(fig_line, use_container_width=True, key=f"line_{titulo_contexto}")

    # --- DETALHAMENTO ---
    st.markdown("### 📋 Detalhamento")
    cols_show = ['Data', 'Cliente_Final', 'Contato', 'Atendido por', 'Status']
    st.dataframe(df_ranking[[c for c in cols_show if c in df_ranking.columns]], use_container_width=True,
                 hide_index=True)

    # --- OPÇÕES DE ADMINISTRADOR (MANTIDAS) ---
    is_local_file = caminho_arquivo_atual and not str(caminho_arquivo_atual).startswith("BQ:")

    if st.session_state.get('user_role') == 'admin':
        st.markdown("<br>", unsafe_allow_html=True)

        with st.expander("🛠️ Opções de Administrador"):

            # === LABEL DO MÊS (disponível para TODOS os meses) ===
            st.markdown("##### 📌 Label do Mês")
            st.caption("Adicione um aviso visível no topo da aba (ex: PARCIAL - 09/02)")

            # Extrai mes_key do caminho
            mes_key_admin = ""
            if caminho_arquivo_atual:
                if str(caminho_arquivo_atual).startswith("BQ:"):
                    # BQ:2026_02 -> 2026_02
                    mes_key_admin = str(caminho_arquivo_atual).replace("BQ:", "")
                else:
                    # data/2026_02.csv -> 2026_02
                    nome_base = os.path.basename(str(caminho_arquivo_atual)).replace(".csv", "")
                    partes = nome_base.split('_', 1)
                    if len(partes) >= 2:
                        mes_limpo = partes[1].split(' ')[0]
                        mes_key_admin = f"{partes[0]}_{mes_limpo}"

            if mes_key_admin:
                # Carrega label atual
                labels_atual = load_month_labels()
                label_existente = labels_atual.get(mes_key_admin, "")

                col_lbl1, col_lbl2, col_lbl3 = st.columns([3, 1, 1])
                with col_lbl1:
                    novo_label = st.text_input(
                        "Label", value=label_existente,
                        placeholder="Ex: PARCIAL - 09/02",
                        key=f"lbl_in_{titulo_contexto}",
                        label_visibility="collapsed"
                    )
                with col_lbl2:
                    if st.button("💾 Salvar", key=f"lbl_save_{titulo_contexto}"):
                        if novo_label:
                            ok, msg = save_month_label(mes_key_admin, novo_label)
                            if ok:
                                st.success(f"Label salvo: {novo_label}")
                                st.rerun()
                            else:
                                st.error(f"Erro BigQuery: {msg}")
                        else:
                            st.warning("Digite um label.")
                with col_lbl3:
                    if st.button("🗑️ Limpar", key=f"lbl_clear_{titulo_contexto}"):
                        ok, msg = save_month_label(mes_key_admin, "")
                        if ok:
                            st.success("Label removido!")
                            st.rerun()
                        else:
                            st.error(f"Erro BigQuery: {msg}")
            else:
                st.info("Não foi possível identificar a chave do mês.")

            # === OPÇÕES DE ARQUIVO LOCAL (somente para arquivos locais) ===
            if is_local_file:
                st.markdown("---")

                # 1. RENOMEAR
                st.markdown("##### ✏️ Renomear Mês")
                col_ren1, col_ren2 = st.columns([3, 1])
                with col_ren1:
                    nome_atual_arquivo = os.path.basename(caminho_arquivo_atual).replace(".csv", "")
                    novo_nome = st.text_input("Novo nome do arquivo:", value=nome_atual_arquivo,
                                              key=f"rn_in_{titulo_contexto}", label_visibility="collapsed")
                with col_ren2:
                    if st.button("Renomear", key=f"rn_bt_{titulo_contexto}"):
                        if novo_nome and novo_nome != nome_atual_arquivo:
                            novo_caminho = os.path.join("data", novo_nome + ".csv")
                            try:
                                os.rename(caminho_arquivo_atual, novo_caminho)
                                st.success("Renomeado!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro: {e}")

                st.markdown("---")

                # 2. SUBSTITUIR
                st.markdown("##### 📂 Substituir Arquivo (CSV)")
                col_up1, col_up2 = st.columns([3, 1])
                with col_up1:
                    up_sub = st.file_uploader("Upload do novo arquivo", type=['csv'], key=f"up_sub_{titulo_contexto}",
                                              label_visibility="collapsed")
                with col_up2:
                    if up_sub and st.button("Substituir", key=f"bt_sub_{titulo_contexto}"):
                        with open(caminho_arquivo_atual, "wb") as f:
                            f.write(up_sub.getbuffer())
                        st.success("Atualizado!")
                        st.rerun()

                st.markdown("---")

                # 3. EXCLUIR
                st.markdown("##### 🗑️ Zona de Perigo")
                col_del1, col_del2 = st.columns([3, 1])
                with col_del1:
                    check_del = st.checkbox("Confirmar exclusão definitiva deste mês", key=f"chk_del_{titulo_contexto}")
                with col_del2:
                    if st.button("Excluir Mês", type="primary", key=f"bt_del_{titulo_contexto}"):
                        if check_del:
                            try:
                                os.remove(caminho_arquivo_atual)
                                st.success("Excluído!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro: {e}")
                        else:
                            st.warning("Marque a caixa ao lado.")


def main_app():
    """Aplicação principal."""
    
    tv_mode = st.session_state.get('tv_mode', False)
    dark_mode = st.session_state.get('dark_mode', False)
    
    # Aplica estilos CSS baseados no modo
    st.markdown(get_css_styles(tv_mode, dark_mode), unsafe_allow_html=True)
    
    # Header
    render_header(tv_mode)
    
    # --- LÓGICA DE EXIBIÇÃO ---
    if st.session_state.get('show_config', False) and st.session_state['user_role'] == 'admin':
        # PAINEL DE CONFIGURAÇÕES (MANTIDO)
        st.markdown("### ⚙️ Painel de Configurações")
        if st.button("⬅️ Voltar"): 
            st.session_state['show_config'] = False
            st.rerun()
        
        if 'feedback_sucesso' in st.session_state: 
            st.success(st.session_state['feedback_sucesso'])
            del st.session_state['feedback_sucesso']

        c1, c2 = st.columns(2)
        with c1:
            st.info("**📁 Iniciar um Novo Mês (Arquivo Local)**")
            nm = st.text_input("Nome (ex: 2026_03)", key="novo_mes_nome")
            up = st.file_uploader("CSV Inicial", key="novo_mes_arq")
            if st.button("➕ Criar Mês"):
                if nm and up:
                    if not nm.endswith(".csv"): nm += ".csv"
                    if not os.path.exists("data"): os.makedirs("data")
                    with open(os.path.join("data", nm), "wb") as f:
                        f.write(up.getbuffer())
                    st.session_state['feedback_sucesso'] = f"Mês {nm} criado com sucesso!"
                    st.rerun()
        with c2:
            st.warning("**📊 Atualizar Base Global**")
            up_xlsx = st.file_uploader("statusContatos.xlsx", type=['xlsx'], key="global_xlsx")
            if up_xlsx and st.button("💾 Salvar Base"):
                with open("statusContatos.xlsx", "wb") as f: 
                    f.write(up_xlsx.getbuffer())
                st.success("Atualizado!")

    else:
        # TELA DE DASHBOARD
        # 1. Carrega estrutura UMA VEZ para usar em tudo
        dados_anos = listar_arquivos_por_ano()
        lista_anos = list(dados_anos.keys())

        if not lista_anos:
            render_empty_state(
                "Nenhum dado disponível",
                icon="📂",
                suggestion="Faça upload de arquivos CSV ou configure a conexão com BigQuery",
                tv_mode=tv_mode
            )
            return

        abas_topo_nomes = lista_anos + ["📈 Análise"]
        abas_topo = st.tabs(abas_topo_nomes)

        # Carrega labels uma vez para usar em todas as abas
        month_labels = load_month_labels()

        # === 1. LOOP DOS ANOS (VISUALIZAÇÃO MENSAL) ===
        for i, ano in enumerate(lista_anos):
            with abas_topo[i]:
                meses_do_ano = dados_anos[ano]

                # Monta nomes das abas COM indicador de label
                nomes_meses = []
                for m in meses_do_ano:
                    nome = m['display']
                    mes_key = f"{ano}_{str(m['mes_raw']).zfill(2)}"
                    lbl = month_labels.get(mes_key, "")
                    if lbl:
                        tag_curta = lbl.split(" - ")[0].strip()
                        nome += f" | {tag_curta}"
                    nomes_meses.append(nome)

                if not nomes_meses:
                    render_empty_state(
                        f"Sem dados para {ano}",
                        icon="📅",
                        suggestion="Adicione arquivos CSV para este ano",
                        tv_mode=tv_mode
                    )
                else:
                    abas_meses = st.tabs(nomes_meses)
                    for j, aba_mes_nome in enumerate(nomes_meses):
                        with abas_meses[j]:
                            caminho = meses_do_ano[j]['caminho']
                            mes_key = f"{ano}_{str(meses_do_ano[j]['mes_raw']).zfill(2)}"

                            # Badge amarelo se houver label
                            lbl = month_labels.get(mes_key, "")
                            if lbl:
                                st.markdown(f"""
                                    <div style="
                                        background: linear-gradient(135deg, #FFA726, #FB8C00);
                                        color: white;
                                        padding: 0.5rem 1rem;
                                        border-radius: 10px;
                                        font-weight: 700;
                                        font-size: 0.9rem;
                                        text-align: center;
                                        margin-bottom: 0.75rem;
                                        box-shadow: 0 2px 8px rgba(251, 140, 0, 0.3);
                                    ">
                                        ⚠️ {lbl}
                                    </div>
                                """, unsafe_allow_html=True)

                            with st.spinner("Carregando dados..."):
                                df = carregar_dados_mes(caminho)
                            
                            if df is not None:
                                renderizar_metricas_limpas(df, f"{ano}_{j}", caminho_arquivo_atual=caminho)
                            else:
                                render_empty_state(
                                    "Erro ao carregar dados",
                                    icon="⚠️",
                                    suggestion="Verifique se o arquivo existe e está no formato correto",
                                    tv_mode=tv_mode
                                )

        # === 2. ABA PERSONALIZADA (ANÁLISE DE PERÍODO) ===
        with abas_topo[-1]:
            st.markdown("### 🔍 Análise de Período Personalizado")
            
            # QUICK FILTERS COM BOTÕES (PILLS STYLE)
            st.markdown("#### ⚡ Período Rápido")
            c_quick = st.columns(5)
            
            periodo_options = {
                "3 Meses": ("Últimos 3 Meses", 90),
                "6 Meses": ("Últimos 6 Meses", 180),
                "Ano 2026": ("Ano Atual (2026)", "year"),
                "Histórico": ("Todo o Período", "all"),
                "Personalizado": ("Selecionar Datas...", "custom")
            }
            
            # Initialize period in session state if not exists
            if 'selected_period' not in st.session_state:
                st.session_state['selected_period'] = "3 Meses"
            
            for idx, (label, (full_name, value)) in enumerate(periodo_options.items()):
                with c_quick[idx]:
                    if st.button(
                        label, 
                        key=f"btn_period_{label}",
                        use_container_width=True,
                        type="primary" if st.session_state.get('selected_period') == label else "secondary"
                    ):
                        st.session_state['selected_period'] = label
                        st.rerun()
            
            st.markdown("<div style='margin-top: 1rem;'></div>", unsafe_allow_html=True)
            
            hoje = datetime.now().date()
            d_inicio, d_fim = None, None
            
            selected = st.session_state.get('selected_period', "3 Meses")
            
            if selected == "3 Meses":
                d_inicio = hoje - timedelta(days=90)
                d_fim = hoje
            elif selected == "6 Meses":
                d_inicio = hoje - timedelta(days=180)
                d_fim = hoje
            elif selected == "Ano 2026":
                d_inicio = datetime(2026, 1, 1).date()
                d_fim = datetime(2026, 12, 31).date()
            elif selected == "Histórico":
                d_inicio = datetime(2020, 1, 1).date()
                d_fim = hoje
            elif selected == "Personalizado":
                c_date1, c_date2 = st.columns(2)
                with c_date1:
                    d_inicio = st.date_input("📅 Data Inicial", hoje - timedelta(days=30), key="date_start_custom")
                with c_date2:
                    d_fim = st.date_input("📅 Data Final", hoje, key="date_end_custom")
            
            if d_inicio and d_fim:
                st.info(f"📆 Período selecionado: **{d_inicio.strftime('%d/%m/%Y')}** até **{d_fim.strftime('%d/%m/%Y')}**")
            
            if st.button("🔎 Pesquisar Dados", type="primary", key="btn_analise_loop", use_container_width=True):
                if not d_inicio or not d_fim:
                    st.warning("⚠️ Selecione um período válido.")
                else:
                    lista_dfs = []

                    with st.spinner(f"🔄 Buscando dados entre {d_inicio.strftime('%d/%m/%Y')} e {d_fim.strftime('%d/%m/%Y')}..."):
                        for ano_key in dados_anos:
                            for item_mes in dados_anos[ano_key]:
                                identificador = item_mes['caminho']
                                df_temp = carregar_dados_mes(identificador)

                                if df_temp is not None and not df_temp.empty:
                                    min_d = df_temp['Dia'].min()
                                    max_d = df_temp['Dia'].max()

                                    if not (max_d < d_inicio or min_d > d_fim):
                                        lista_dfs.append(df_temp)

                        # Consolidação
                        if lista_dfs:
                            df_full = pd.concat(lista_dfs, ignore_index=True)

                            # Filtro Fino de Data
                            df_full['Dia'] = pd.to_datetime(df_full['Data']).dt.date
                            df_filtrado = df_full[(df_full['Dia'] >= d_inicio) & (df_full['Dia'] <= d_fim)]

                            if df_filtrado.empty:
                                render_empty_state(
                                    "Nenhum dado neste período",
                                    icon="📭",
                                    suggestion=f"Não há registros entre {d_inicio.strftime('%d/%m/%Y')} e {d_fim.strftime('%d/%m/%Y')}",
                                    tv_mode=tv_mode
                                )
                            else:
                                st.success(
                                    f"✅ Consolidado com sucesso: **{len(df_filtrado):,}** registros de **{len(lista_dfs)}** tabelas.".replace(",", ".")
                                )
                                renderizar_metricas_limpas(df_filtrado, "analise_loop_v1")
                        else:
                            render_empty_state(
                                "Nenhuma tabela encontrada",
                                icon="🔍",
                                suggestion="Não há dados disponíveis para o período selecionado",
                                tv_mode=tv_mode
                            )


# --- EXECUÇÃO ---
if __name__ == "__main__":
    if autenticar():
        main_app()